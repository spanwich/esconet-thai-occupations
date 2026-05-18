#!/usr/bin/env python3
"""
Build CHANCEDEE V3 Job Industry Master from TSIC Pass 3 + Industry_Xwalk.
Outputs:
  - job_industry_master.parquet   (programmatic use)
  - chancedee_job_industry_master.xlsx (review / deck)
"""
import duckdb, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO = "/tmp/esconet-thai"
OUT = "/home/claude"

con = duckdb.connect()

# ---- name trust tier ----
def trust_tier(q):
    if q in ("clean_authoritative",): return "highest"
    if q in ("clean", "llm_clean"):   return "high"
    if q in ("mostly_clean",):        return "medium"
    if q in ("llm_review",):          return "review"
    if q in ("partial_corruption", "heavy_corruption", "llm_uncertain"): return "low"
    if q in ("no_name",):             return "none"
    return "unknown"

# ---- Section letter for each Activity/Class/Group/Division ----
# Section codes are letters A-U; Division/Group/Class/Activity are digits.
# Section of a code is determined by its Division (first 2 digits) per TSIC structure.
section_div_ranges = {
    "A":("01","03"),"B":("05","09"),"C":("10","33"),"D":("35","35"),"E":("36","39"),
    "F":("41","43"),"G":("45","47"),"H":("49","53"),"I":("55","56"),"J":("58","63"),
    "K":("64","66"),"L":("68","68"),"M":("69","75"),"N":("77","82"),"O":("84","84"),
    "P":("85","85"),"Q":("86","88"),"R":("90","93"),"S":("94","96"),"T":("97","98"),"U":("99","99"),
}
def section_of(code, level):
    if level == "Section": return code if (isinstance(code,str) and code.isalpha()) else None
    if not isinstance(code,str) or len(code) < 2: return None
    div = code[:2]
    if not div.isdigit(): return None
    for sec, (lo, hi) in section_div_ranges.items():
        if lo <= div <= hi: return sec
    return None

# ---- Pull TSIC + name fields ----
tsic = con.execute(f"""
SELECT code, level, name_th, name_en, parent_code,
       name_quality, name_quality_score, llm_confidence
FROM '{REPO}/THAI/tsic_pass3.parquet'
ORDER BY code
""").fetchdf()
tsic["name_trust"] = tsic["name_quality"].map(trust_tier)
tsic["section_code"] = tsic.apply(lambda r: section_of(r["code"], r["level"]), axis=1)

# ---- Section labels ----
sec_labels_th = {
  "A":"เกษตรกรรม การป่าไม้ และการประมง", "B":"การทำเหมืองแร่และเหมืองหิน",
  "C":"การผลิต", "D":"ไฟฟ้า ก๊าซ ไอน้ำ และระบบปรับอากาศ",
  "E":"การจัดหาน้ำ การจัดการน้ำเสียและของเสีย และการบำบัด", "F":"การก่อสร้าง",
  "G":"การขายส่งและการขายปลีก การซ่อมยานยนต์และจักรยานยนต์", "H":"การขนส่งและสถานที่เก็บสินค้า",
  "I":"ที่พักแรมและบริการด้านอาหาร", "J":"ข้อมูลข่าวสารและการสื่อสาร",
  "K":"กิจกรรมทางการเงินและการประกันภัย", "L":"กิจกรรมอสังหาริมทรัพย์",
  "M":"กิจกรรมทางวิชาชีพ วิทยาศาสตร์ และเทคนิค", "N":"กิจกรรมการบริหารและบริการสนับสนุน",
  "O":"การบริหารราชการ การป้องกันประเทศ และการประกันสังคมภาคบังคับ", "P":"การศึกษา",
  "Q":"กิจกรรมด้านสุขภาพและงานสังคมสงเคราะห์", "R":"ศิลปะ ความบันเทิง และนันทนาการ",
  "S":"กิจกรรมบริการอื่น ๆ", "T":"กิจกรรมการจ้างงานในครัวเรือนส่วนบุคคล",
  "U":"กิจกรรมขององค์การระหว่างประเทศและภาคีสมาชิก",
}
tsic["section_name_th"] = tsic["section_code"].map(sec_labels_th)

# ---- Hierarchy path (Section › Division › Group › Class › Activity) ----
by_code = tsic.set_index("code")[["level","name_th","parent_code"]].to_dict("index")
def path_codes(code):
    chain = []; cur = code; seen = set()
    while cur and cur in by_code and cur not in seen:
        seen.add(cur); chain.append(cur)
        cur = by_code[cur]["parent_code"]
    return list(reversed(chain))
def path_str(code):
    return " › ".join(path_codes(code))
tsic["path_codes"] = tsic["code"].map(path_codes)
tsic["path_str"] = tsic["code"].map(path_str)
tsic["depth"] = tsic["path_codes"].str.len()

# ---- Crosswalk reach + occupation summary ----
# Aggregate per TSIC: occ count at three thresholds, top occupations.
flat = con.execute(f"""
SELECT tsic_code, soc_code, soc_title, onet_soc_code, onet_soc_title, emp_2024_k
FROM '{REPO}/Industry_Xwalk/tsic_occupation_flat.parquet'
WHERE onet_soc_code IS NOT NULL
""").fetchdf()

# Distinct soc_code per tsic (drop dup O*NET sub-codes) for counts; for top list keep first occurrence.
soc_per_tsic = (flat[["tsic_code","soc_code","emp_2024_k"]]
                .drop_duplicates(subset=["tsic_code","soc_code"]))
def topN(group, n=8):
    return list(group.sort_values("emp_2024_k", ascending=False).head(n)["soc_title_emp"])

# Build a label "Title (xxxk)" string
soc_titles = (flat.drop_duplicates(subset=["tsic_code","soc_code"])
                  .assign(soc_title_emp=lambda d: d["soc_title"]+" ("+d["emp_2024_k"].round(0).astype(int).astype(str)+"k)"))

top_by_tsic = (soc_titles.groupby("tsic_code")
               .apply(lambda g: " | ".join(g.sort_values("emp_2024_k", ascending=False).head(8)["soc_title_emp"]))
               .rename("top_occupations").reset_index())

agg = (soc_per_tsic.groupby("tsic_code")
       .agg(n_occ_all=("soc_code","nunique"),
            n_occ_ge_1k=("emp_2024_k", lambda s: int((s>=1).sum())),
            n_occ_ge_10k=("emp_2024_k", lambda s: int((s>=10).sum())),
            n_occ_ge_50k=("emp_2024_k", lambda s: int((s>=50).sum())))
       .reset_index())
agg = agg.merge(top_by_tsic, on="tsic_code", how="left")

# ---- Join ----
master = tsic.merge(agg, left_on="code", right_on="tsic_code", how="left").drop(columns=["tsic_code"])
master["has_occupations"] = master["n_occ_all"].notna()
for c in ["n_occ_all","n_occ_ge_1k","n_occ_ge_10k","n_occ_ge_50k"]:
    master[c] = master[c].fillna(0).astype(int)

# ---- Usability flag ----
# usable_in_picker = high-trust name + has crosswalk to occupations
master["usable_in_picker"] = master["name_trust"].isin(["highest","high"]) & master["has_occupations"]

# ---- Column order ----
cols = ["code","level","depth","section_code","section_name_th",
        "name_th","name_en","parent_code","path_str",
        "name_trust","name_quality","llm_confidence",
        "has_occupations","n_occ_all","n_occ_ge_1k","n_occ_ge_10k","n_occ_ge_50k",
        "top_occupations","usable_in_picker"]
master = master[cols].sort_values(["section_code","code"], na_position="last")

# ---- Save parquet ----
master.to_parquet(f"{OUT}/job_industry_master.parquet", index=False)
print(f"master rows: {len(master)}")
print(master["name_trust"].value_counts())
print(master["usable_in_picker"].value_counts())

# ---- Filter strength analysis table ----
filt = []
universe = flat["soc_code"].nunique()
for thr, label in [(0,"all (no threshold)"),(1,"≥ 1k workers"),(5,"≥ 5k workers"),
                   (10,"≥ 10k workers"),(50,"≥ 50k workers")]:
    sub = soc_per_tsic[soc_per_tsic["emp_2024_k"]>=thr]
    avg_occ = sub.groupby("tsic_code")["soc_code"].nunique().mean()
    pct = (avg_occ/universe)*100 if avg_occ==avg_occ else 0
    filt.append(dict(threshold=label, avg_occupations_per_tsic=round(avg_occ,1),
                     pct_of_universe=round(pct,1), interpretation=""))
filt_df = pd.DataFrame(filt)
filt_df.loc[filt_df["threshold"]=="all (no threshold)","interpretation"] = "weak filter — keeps ~1/4 of universe"
filt_df.loc[filt_df["threshold"]=="≥ 1k workers","interpretation"]    = "drops marginal occupations"
filt_df.loc[filt_df["threshold"]=="≥ 5k workers","interpretation"]    = "medium filter"
filt_df.loc[filt_df["threshold"]=="≥ 10k workers","interpretation"]    = "strong filter — recommended default"
filt_df.loc[filt_df["threshold"]=="≥ 50k workers","interpretation"]    = "very narrow — only signature occupations"
filt_df["universe_size"] = universe
print("\nfilter strength table:")
print(filt_df.to_string(index=False))

# ---- Build xlsx ----
wb = Workbook()

# ---------------- README ----------------
ws = wb.active; ws.title = "README"
readme = [
  ("CHANCEDEE V3 — Job Industry Master (TSIC-derived)",""),
  ("",""),
  ("Source",  "github.com/spanwich/esconet-thai-occupations"),
  ("Built from", "THAI/tsic_pass3.parquet + Industry_Xwalk/tsic_occupation_flat.parquet"),
  ("Built date", "2026-05-15"),
  ("",""),
  ("What this is", ""),
  ("", "ลิสต์อุตสาหกรรมฝั่งไทย (Job Industry THAI ในไดอะแกรม) ที่ derive จาก TSIC Pass 3"),
  ("", "พร้อม trust tag, hierarchy path, และจำนวน occupation ที่ reach ผ่าน crosswalk"),
  ("", "→ ESCONET (TSIC → ISIC4 → NAICS3 → BLS NEM → O*NET-SOC)"),
  ("",""),
  ("Sheets", ""),
  ("Master", "1,892 rows — all TSIC codes with trust tag + occ reach + top occupations"),
  ("Filter_Strength", "How well industry filters occupation, at different employment thresholds"),
  ("Usable_in_Picker", "Subset where name_trust ∈ (highest,high) AND has_occupations — UI-ready"),
  ("Excluded", "Codes excluded from Usable_in_Picker, with reason"),
  ("Sections", "21 TSIC Sections with Thai labels"),
  ("Legend", "Column definitions, trust tiers, level codes"),
  ("",""),
  ("Filter strength — the question Ford asked", ""),
  ("", "Crosswalk filter strength depends on employment threshold:"),
  ("", "• No threshold: filter is WEAK (~238 occupations per TSIC, ~29% of 810-SOC universe)"),
  ("", "• emp ≥ 10k workers: filter is STRONG (~39 per TSIC, ~5% of universe) — recommended"),
  ("", "• emp ≥ 50k workers: filter is VERY NARROW (~12 per TSIC) — signature occupations only"),
  ("", "Universe = 810 distinct SOC codes (parent occupations). Each SOC may map to multiple O*NET-SOC specializations."),
  ("",""),
  ("Key caveat", ""),
  ("", "emp_2024_k = BLS US employment (thousands). Use as RELATIVE ranking signal,"),
  ("", "NOT as a count of Thai workers. Thai staffing structure differs from US."),
  ("",""),
  ("How to use", ""),
  ("Picker UI", "Filter Master where usable_in_picker=TRUE, group by section_code"),
  ("Industry filter", "Pick a code; reachable occupations = top_occupations column"),
  ("Programmatic", "Load job_industry_master.parquet and join with tsic_occupation_flat.parquet"),
]
for i,(k,v) in enumerate(readme, start=1):
    ws[f"A{i}"]=k; ws[f"B{i}"]=v
ws.column_dimensions["A"].width=24; ws.column_dimensions["B"].width=98
ws["A1"].font = Font(bold=True, size=14)
for row in (7,12,20,28,32):
    ws[f"A{row}"].font = Font(bold=True, size=11)

# ---------------- Master ----------------
ws = wb.create_sheet("Master")
m = master.copy()
ws.append(list(m.columns))
for r in m.itertuples(index=False):
    ws.append(list(r))
for c, col in enumerate(m.columns, start=1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="305496")
    cell.alignment = Alignment(horizontal="center", vertical="center")
widths = {"code":10,"level":10,"depth":6,"section_code":8,"section_name_th":40,
          "name_th":48,"name_en":48,"parent_code":12,"path_str":28,
          "name_trust":10,"name_quality":18,"llm_confidence":12,
          "has_occupations":12,"n_occ_all":10,"n_occ_ge_1k":12,"n_occ_ge_10k":13,"n_occ_ge_50k":13,
          "top_occupations":90,"usable_in_picker":14}
for c, col in enumerate(m.columns, start=1):
    ws.column_dimensions[get_column_letter(c)].width = widths.get(col, 14)
ws.freeze_panes = "C2"
ws.auto_filter.ref = ws.dimensions

# ---------------- Filter_Strength ----------------
ws = wb.create_sheet("Filter_Strength")
ws.append(["Employment threshold","Avg occupations per TSIC","% of 810-SOC universe","Interpretation","Universe size"])
for r in filt_df.itertuples(index=False):
    ws.append(list(r))
ws.append([])
ws.append(["Notes",""])
ws.append(["","Universe = 810 distinct SOC parent occupations reachable through crosswalk"])
ws.append(["","(each SOC may map to multiple O*NET-SOC specializations — 953 total O*NET-SOC codes)"])
ws.append(["","Average is over the 1,780 TSIC codes that have any occupation reach"])
ws.append(["","Recommended default: emp ≥ 10k for picker; emp ≥ 50k for 'characteristic occupations'"])
for c in range(1,6):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="305496")
for col, w in [("A",24),("B",26),("C",22),("D",46),("E",14)]:
    ws.column_dimensions[col].width = w

# ---------------- Usable_in_Picker ----------------
ws = wb.create_sheet("Usable_in_Picker")
sub = master[master["usable_in_picker"]==True].copy()
keep_cols = ["section_code","section_name_th","level","code","name_th","name_en",
             "path_str","n_occ_ge_10k","top_occupations"]
sub = sub[keep_cols]
ws.append(list(sub.columns))
for r in sub.itertuples(index=False):
    ws.append(list(r))
for c in range(1, len(sub.columns)+1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="548235")
widths2 = {"section_code":8,"section_name_th":40,"level":10,"code":10,
           "name_th":48,"name_en":48,"path_str":28,"n_occ_ge_10k":13,"top_occupations":90}
for c, col in enumerate(sub.columns, start=1):
    ws.column_dimensions[get_column_letter(c)].width = widths2.get(col, 14)
ws.freeze_panes = "C2"
ws.auto_filter.ref = ws.dimensions
print(f"Usable_in_Picker rows: {len(sub)}")

# ---------------- Excluded ----------------
ws = wb.create_sheet("Excluded")
exc = master[master["usable_in_picker"]==False].copy()
def excl_reason(r):
    if r["name_trust"]=="none":   return "no_name (parent code without label)"
    if r["name_trust"]=="low":    return "name corrupted (low trust)"
    if r["name_trust"]=="review": return "name needs human review (LLM low confidence)"
    if r["name_trust"]=="medium": return "name partly clean (medium trust)"
    if not r["has_occupations"]:  return "no crosswalk to occupations (Thailand-specific / unmatched ISIC)"
    return "other"
exc["exclusion_reason"] = exc.apply(excl_reason, axis=1)
exc_cols = ["code","level","section_code","name_th","name_trust","has_occupations","exclusion_reason"]
exc = exc[exc_cols]
ws.append(list(exc.columns))
for r in exc.itertuples(index=False):
    ws.append(list(r))
for c in range(1, len(exc.columns)+1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="C00000")
widths3 = {"code":10,"level":10,"section_code":8,"name_th":48,
           "name_trust":10,"has_occupations":14,"exclusion_reason":56}
for c, col in enumerate(exc.columns, start=1):
    ws.column_dimensions[get_column_letter(c)].width = widths3.get(col, 14)
ws.freeze_panes = "C2"
ws.auto_filter.ref = ws.dimensions
print(f"Excluded rows: {len(exc)}")

# ---------------- Sections ----------------
ws = wb.create_sheet("Sections")
ws.append(["section_code","section_name_th","n_codes_in_section"])
sec_counts = master.groupby("section_code").size().reset_index(name="n")
for code, lbl in sec_labels_th.items():
    n = int(sec_counts[sec_counts["section_code"]==code]["n"].sum())
    ws.append([code, lbl, n])
for c in range(1,4):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="305496")
for col,w in [("A",14),("B",60),("C",22)]:
    ws.column_dimensions[col].width = w

# ---------------- Legend ----------------
ws = wb.create_sheet("Legend")
legend = [
 ("Column","Meaning"),
 ("code","TSIC code (1, 2, 3, 4, or 5 digits — Section is alpha)"),
 ("level","Section | Division | Group | Class | Activity"),
 ("depth","Depth in TSIC tree (1=Section ... 5=Activity)"),
 ("section_code","TSIC Section letter (A-U)"),
 ("section_name_th","Thai label of the Section"),
 ("name_th","Cleaned Thai name (see TSIC_README.md for cleanup history)"),
 ("name_en","English name where available (sourced from upstream)"),
 ("parent_code","Parent TSIC code (NULL for Section)"),
 ("path_str","Breadcrumb of codes from Section to this code"),
 ("name_trust","Trust tier: highest | high | medium | review | low | none"),
 ("name_quality","Raw TSIC Pass 3 quality flag (clean / llm_clean / no_name / ...)"),
 ("llm_confidence","If LLM-corrected: confidence ∈ [0,1]; else null"),
 ("has_occupations","TRUE if reachable through crosswalk to O*NET-SOC occupations"),
 ("n_occ_all","Distinct O*NET-SOC occupations reachable, no employment threshold"),
 ("n_occ_ge_1k","... where US BLS emp ≥ 1k workers (drops marginal)"),
 ("n_occ_ge_10k","... where US BLS emp ≥ 10k workers (recommended default filter)"),
 ("n_occ_ge_50k","... where US BLS emp ≥ 50k workers (signature occupations)"),
 ("top_occupations","Top 8 occupations by emp_2024_k, pipe-separated 'Title (Nk)'"),
 ("usable_in_picker","TRUE if name_trust ∈ (highest,high) AND has_occupations"),
 ("",""),
 ("Trust tier definitions",""),
 ("highest","DBD-authoritative (157 records, gold standard)"),
 ("high","Dict-validated 'clean' OR LLM corrected with conf ≥ 0.9"),
 ("medium","'mostly_clean' — dict score 0.70-0.95"),
 ("review","LLM corrected with conf < 0.9 — human review recommended"),
 ("low","Partial or heavy corruption, or LLM uncertain"),
 ("none","Empty name (typically parent codes Division/Group/Class)"),
]
for r in legend:
    ws.append(list(r))
ws["A1"].font = Font(bold=True); ws["B1"].font = Font(bold=True)
ws["A1"].fill = PatternFill("solid", start_color="305496"); ws["A1"].font = Font(bold=True, color="FFFFFF")
ws["B1"].fill = PatternFill("solid", start_color="305496"); ws["B1"].font = Font(bold=True, color="FFFFFF")
ws["A22"].font = Font(bold=True, size=11)
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 90

wb.save(f"{OUT}/chancedee_job_industry_master.xlsx")
print(f"\nWrote {OUT}/chancedee_job_industry_master.xlsx")
print(f"Wrote {OUT}/job_industry_master.parquet")
